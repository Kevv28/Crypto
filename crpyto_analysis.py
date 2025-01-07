import requests
import pandas as pd
import schedule
import time
from openpyxl import Workbook
from fpdf import FPDF
from datetime import datetime

API_URL = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest"
API_KEY = "dfdfbe45-ddd2-4e3e-940d-3d5b0c5ab719"  
HEADERS = {"X-CMC_PRO_API_KEY": API_KEY}

def fetch_crypto_data():
    params = {"start": "1", "limit": "50", "convert": "USD"}
    response = requests.get(API_URL, headers=HEADERS, params=params)
    data = response.json()
    
    
    cryptocurrencies = []
    for item in data["data"]:
        cryptocurrencies.append({
            "Name": item["name"],
            "Symbol": item["symbol"],
            "Price (USD)": round(item["quote"]["USD"]["price"], 2),
            "Market Cap": round(item["quote"]["USD"]["market_cap"], 2),
            "24h Volume": round(item["quote"]["USD"]["volume_24h"], 2),
            "24h Change (%)": round(item["quote"]["USD"]["percent_change_24h"], 2)
        })
    return pd.DataFrame(cryptocurrencies)


def analyze_data(df):
    top_5 = df.nlargest(5, "Market Cap")
    avg_price = df["Price (USD)"].mean()
    highest_change = df.loc[df["24h Change (%)"].idxmax()]
    lowest_change = df.loc[df["24h Change (%)"].idxmin()]
    return top_5, avg_price, highest_change, lowest_change


def update_excel(df):
    with pd.ExcelWriter("crypto_data.xlsx", engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name="Live Data")
    print(f"Excel updated at {datetime.now()}")


def generate_report(df, top_5, avg_price, high_change, low_change):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, txt="Cryptocurrency Market Analysis", ln=True, align="C")
    pdf.ln(10)
    
    
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(5)
    pdf.cell(200, 10, txt="Key Insights:", ln=True)
    pdf.ln(5)
    pdf.cell(200, 10, txt=f"Average Price of Top 50 Cryptocurrencies: ${avg_price:.2f}", ln=True)
    pdf.cell(200, 10, txt=f"Highest 24h Change: {high_change['Name']} ({high_change['24h Change (%)']}%)", ln=True)
    pdf.cell(200, 10, txt=f"Lowest 24h Change: {low_change['Name']} ({low_change['24h Change (%)']}%)", ln=True)
    pdf.ln(10)
    

    pdf.cell(200, 10, txt="Top 5 Cryptocurrencies by Market Cap:", ln=True)
    pdf.ln(5)
    for idx, row in top_5.iterrows():
        pdf.cell(200, 10, txt=f"{row['Name']} ({row['Symbol']}): Market Cap = ${row['Market Cap']:,}", ln=True)
    pdf.ln(10)
    

    pdf_file_name = "crypto_analysis_report.pdf"
    pdf.output(pdf_file_name)
    print(f"Report generated: {pdf_file_name}")


def task():
    print("Fetching live cryptocurrency data...")
    df = fetch_crypto_data()
    print("Analyzing data...")
    top_5, avg_price, high_change, low_change = analyze_data(df)
    print("Updating Excel...")
    update_excel(df)
    print("Generating PDF report...")
    generate_report(df, top_5, avg_price, high_change, low_change)
    print("Task completed.")

task()


schedule.every(5).minutes.do(task)

print("Starting live updates...")
while True:
    schedule.run_pending()
    time.sleep(1)
